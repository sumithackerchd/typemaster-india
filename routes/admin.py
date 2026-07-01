import csv
import io
import json
import os
from types import SimpleNamespace
from datetime import datetime, date, timedelta
from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, jsonify,
    send_file, Response,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from utils.decorators import admin_required
from utils.settings_store import (
    load_settings, save_settings, count_paragraphs, invalidate_paragraph_cache,
    signature_abs_path,
)
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from models import db
from models.user import User
from models.result import Result
from models.download import Download
from utils.extract import (
    extract_text,
    split_paragraphs,
    ExtractionError,
    OCRUnavailable,
    SUPPORTED_EXTS,
    ocr_available,
)

admin = Blueprint(
    "admin",
    __name__,
    url_prefix="/admin"
)


@admin.before_request
def _guard_admin_area():
    """Defense-in-depth: block every /admin/* request for non-admins.

    Even if an individual view is ever added without the ``@admin_required``
    decorator, this guard ensures the entire blueprint stays admin-only.
    """
    from flask import abort
    if not current_user.is_authenticated:
        return redirect(url_for("auth.login", next=request.path))
    if not getattr(current_user, "is_admin", False):
        wants_json = (
            request.is_json
            or request.headers.get("X-Requested-With") == "XMLHttpRequest"
            or request.accept_mimetypes.best == "application/json"
        )
        if wants_json:
            abort(403)
        flash("You are not authorized to access the admin area.", "danger")
        return redirect(url_for("dashboard.dashboard_page"))


BASE_DIR = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
ENGLISH_JSON = os.path.join(BASE_DIR, "static", "data", "paragraphs.json")
HINDI_JSON = os.path.join(BASE_DIR, "static", "data", "hindi_mangal.json")
SETTINGS_JSON = os.path.join(BASE_DIR, "static", "data", "settings.json")


def load_json(path, default):

    if not os.path.exists(path):
        return default

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, data):

    os.makedirs(os.path.dirname(path), exist_ok=True)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


@admin.route("/")
@login_required
@admin_required
def dashboard():

    total_users = User.query.count()
    total_tests = Result.query.count()

    today_start = datetime.combine(date.today(), datetime.min.time())

    today_tests = Result.query.filter(
        Result.created_at >= today_start
    ).count()

    admin_count = User.query.filter_by(is_admin=True).count()

    recent_users = (
        User.query.order_by(User.id.desc()).limit(5).all()
    )

    recent_results = (
        Result.query
        .options(joinedload(Result.user))
        .order_by(Result.created_at.desc())
        .limit(6)
        .all()
    )

    activity_labels = []
    activity_counts = []
    for offset in range(6, -1, -1):
        day = date.today() - timedelta(days=offset)
        start = datetime.combine(day, datetime.min.time())
        end = start + timedelta(days=1)
        activity_labels.append(day.strftime("%a"))
        activity_counts.append(
            Result.query.filter(
                Result.created_at >= start,
                Result.created_at < end,
            ).count()
        )

    top_wpm_rows = (
        db.session.query(
            User.username,
            func.avg(Result.net_wpm).label("avg_wpm"),
        )
        .join(Result, Result.user_id == User.id)
        .group_by(User.id)
        .order_by(func.avg(Result.net_wpm).desc())
        .limit(5)
        .all()
    )
    wpm_labels = [r.username for r in top_wpm_rows]
    wpm_values = [round(float(r.avg_wpm or 0), 1) for r in top_wpm_rows]

    top_active = (
        db.session.query(User, func.count(Result.id).label("test_count"))
        .outerjoin(Result, Result.user_id == User.id)
        .group_by(User.id)
        .order_by(func.count(Result.id).desc())
        .limit(5)
        .all()
    )

    latest_certs = (
        Result.query
        .filter(Result.certificate_id.isnot(None))
        .options(joinedload(Result.user))
        .order_by(Result.created_at.desc())
        .limit(5)
        .all()
    )

    cert_count = Result.query.filter(Result.certificate_id.isnot(None)).count()
    download_items = Download.query.count()
    download_hits = db.session.query(func.sum(Download.download_count)).scalar() or 0
    paragraph_count = count_paragraphs()

    upload_dir = os.path.join(BASE_DIR, "static", "uploads", "downloads")
    storage_bytes = 0
    if os.path.isdir(upload_dir):
        for root, _dirs, files in os.walk(upload_dir):
            for fname in files:
                try:
                    storage_bytes += os.path.getsize(os.path.join(root, fname))
                except OSError:
                    pass

    week_start = datetime.combine(date.today() - timedelta(days=7), datetime.min.time())
    new_users_week = User.query.filter(User.created_at >= week_start).count()

    return render_template(
        "admin/dashboard.html",
        total_users=total_users,
        total_tests=total_tests,
        today_tests=today_tests,
        admin_count=admin_count,
        recent_users=recent_users,
        recent_results=recent_results,
        activity_labels=activity_labels,
        activity_counts=activity_counts,
        wpm_labels=wpm_labels,
        wpm_values=wpm_values,
        cert_count=cert_count,
        download_items=download_items,
        download_hits=int(download_hits),
        paragraph_count=paragraph_count,
        storage_bytes=storage_bytes,
        top_active=top_active,
        latest_certs=latest_certs,
        new_users_week=new_users_week,
        system_ok=True,
    )


@admin.route("/users")
@login_required
@admin_required
def users():

    search = request.args.get("search", "")

    if search:
        users_list = User.query.filter(
            User.full_name.contains(search) |
            User.username.contains(search) |
            User.email.contains(search) |
            User.mobile.contains(search)
        ).order_by(User.id.desc()).all()
    else:
        users_list = User.query.order_by(User.id.desc()).all()

    # Aggregate per-user stats in two grouped queries (avoids N+1).
    tests_map = dict(
        db.session.query(Result.user_id, func.count(Result.id))
        .group_by(Result.user_id)
        .all()
    )
    best_map = dict(
        db.session.query(Result.user_id, func.max(Result.net_wpm))
        .group_by(Result.user_id)
        .all()
    )
    cert_map = dict(
        db.session.query(Result.user_id, func.count(Result.id))
        .filter(Result.certificate_id.isnot(None))
        .group_by(Result.user_id)
        .all()
    )

    stats = {
        u.id: {
            "tests": tests_map.get(u.id, 0),
            "best_wpm": best_map.get(u.id, 0) or 0,
            "certificates": cert_map.get(u.id, 0),
        }
        for u in users_list
    }

    return render_template(
        "admin/user.html",
        users=users_list,
        stats=stats,
        search=search
    )


@admin.route("/users/delete/<int:user_id>")
@login_required
@admin_required
def delete_user(user_id):

    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("admin.users"))

    db.session.delete(user)
    db.session.commit()

    flash("User deleted successfully.", "success")
    return redirect(url_for("admin.users"))


@admin.route("/users/toggle-admin/<int:user_id>")
@login_required
@admin_required
def toggle_admin(user_id):

    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("You cannot change your own admin role.", "danger")
        return redirect(url_for("admin.users"))

    user.is_admin = not user.is_admin
    db.session.commit()

    flash("Admin role updated.", "success")
    return redirect(url_for("admin.users"))


@admin.route("/users/toggle-status/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def toggle_user_status(user_id):
    """Suspend or activate a user account."""
    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("You cannot suspend your own account.", "danger")
        return redirect(request.referrer or url_for("admin.users"))

    user.is_active = not bool(user.is_active)
    db.session.commit()

    if user.is_active:
        flash(f"{user.full_name}'s account has been activated.", "success")
    else:
        flash(f"{user.full_name}'s account has been suspended.", "success")
    return redirect(request.referrer or url_for("admin.users"))


@admin.route("/users/reset-password/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def reset_user_password(user_id):
    """Admin sets a new password for a user."""
    from werkzeug.security import generate_password_hash

    user = User.query.get_or_404(user_id)
    new_password = (request.form.get("new_password") or "").strip()

    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "danger")
        return redirect(request.referrer or url_for("admin.user_profile", user_id=user.id))

    user.password = generate_password_hash(new_password)
    db.session.commit()

    flash(f"Password reset for {user.full_name}.", "success")
    return redirect(request.referrer or url_for("admin.user_profile", user_id=user.id))


@admin.route("/users/<int:user_id>")
@login_required
@admin_required
def user_profile(user_id):
    """Full profile view for a single user with stats and recent activity."""
    user = User.query.get_or_404(user_id)

    results_q = Result.query.filter_by(user_id=user.id)
    total_tests = results_q.count()
    best_wpm = db.session.query(func.max(Result.net_wpm)).filter(
        Result.user_id == user.id
    ).scalar() or 0
    avg_accuracy = db.session.query(func.avg(Result.accuracy)).filter(
        Result.user_id == user.id
    ).scalar() or 0
    certificates = results_q.filter(Result.certificate_id.isnot(None)).count()
    recent_results = results_q.order_by(Result.created_at.desc()).limit(10).all()

    return render_template(
        "admin/user_profile.html",
        user=user,
        total_tests=total_tests,
        best_wpm=int(best_wpm),
        avg_accuracy=round(float(avg_accuracy), 1),
        certificates=certificates,
        recent_results=recent_results,
    )


@admin.route("/results")
@login_required
@admin_required
def results():

    results_list = Result.query.order_by(
        Result.created_at.desc()
    ).all()

    return render_template(
        "admin/result.html",
        results=results_list
    )


@admin.route("/paragraphs/english")
@login_required
@admin_required
def english_paragraphs():

    data = load_json(ENGLISH_JSON, {"english": {"easy": [], "medium": [], "hard": []}})

    return render_template(
        "admin/english_paragraphs.html",
        paragraphs=data.get("english", {}),
        difficulties=["easy", "medium", "hard"],
        ocr_ok=ocr_available(),
    )


@admin.route("/paragraphs/hindi")
@login_required
@admin_required
def hindi_paragraphs():

    data = load_json(HINDI_JSON, {"hindi": {"easy": [], "medium": [], "hard": []}})

    return render_template(
        "admin/hindi_paragraphs.html",
        paragraphs=data.get("hindi", {}),
        difficulties=["easy", "medium", "hard"],
        ocr_ok=ocr_available(),
    )


@admin.route("/paragraphs/save", methods=["POST"])
@login_required
@admin_required
def save_paragraph():

    payload = request.get_json()

    if not payload:
        return jsonify({"success": False, "message": "No data received"}), 400

    language = payload.get("language")
    difficulty = payload.get("difficulty")
    content = (payload.get("content") or "").strip()
    action = payload.get("action", "add")
    index = payload.get("index")

    if language not in ("english", "hindi"):
        return jsonify({"success": False, "message": "Invalid language"}), 400

    if difficulty not in ("easy", "medium", "hard"):
        return jsonify({"success": False, "message": "Invalid difficulty"}), 400

    path = ENGLISH_JSON if language == "english" else HINDI_JSON
    default = {language: {"easy": [], "medium": [], "hard": []}}
    data = load_json(path, default)

    if language not in data:
        data[language] = {"easy": [], "medium": [], "hard": []}

    items = data[language].setdefault(difficulty, [])

    if action == "add":

        if not content:
            return jsonify({"success": False, "message": "Content required"}), 400

        items.append(content)

    elif action == "delete":

        if index is None or index < 0 or index >= len(items):
            return jsonify({"success": False, "message": "Invalid index"}), 400

        items.pop(index)

    elif action == "edit":

        if index is None or index < 0 or index >= len(items):
            return jsonify({"success": False, "message": "Invalid index"}), 400

        if not content:
            return jsonify({"success": False, "message": "Content required"}), 400

        items[index] = content

    else:
        return jsonify({"success": False, "message": "Invalid action"}), 400

    save_json(path, data)
    invalidate_paragraph_cache()

    return jsonify({"success": True, "paragraphs": data[language]})


def _lang_path(language):
    return ENGLISH_JSON if language == "english" else HINDI_JSON


def _load_grouped(language):
    default = {language: {"easy": [], "medium": [], "hard": []}}
    data = load_json(_lang_path(language), default)
    if language not in data:
        data[language] = {"easy": [], "medium": [], "hard": []}
    for d in ("easy", "medium", "hard"):
        data[language].setdefault(d, [])
    return data


@admin.route("/paragraphs/import", methods=["POST"])
@login_required
@admin_required
def import_paragraphs():
    """Bulk import paragraphs from an uploaded file into a difficulty bucket.

    Accepts TXT / DOCX / PDF / CSV / XLSX / JSON (and images when OCR is
    available). The extracted text is split into individual paragraphs,
    de-duplicated against what already exists, and appended.
    """
    language = (request.form.get("language") or "english").lower()
    difficulty = (request.form.get("difficulty") or "medium").lower()

    if language not in ("english", "hindi"):
        return jsonify({"success": False, "message": "Invalid language"}), 400
    if difficulty not in ("easy", "medium", "hard"):
        return jsonify({"success": False, "message": "Invalid difficulty"}), 400

    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"success": False, "message": "No file uploaded."}), 400

    ext = ("." + uploaded.filename.rsplit(".", 1)[-1].lower()) if "." in uploaded.filename else ""
    if ext not in SUPPORTED_EXTS:
        return jsonify({
            "success": False,
            "message": f"Unsupported file type ({ext or 'unknown'}).",
        }), 400

    raw = uploaded.read()
    if not raw:
        return jsonify({"success": False, "message": "The file is empty."}), 400

    try:
        text = extract_text(uploaded.filename, raw, lang=language)
    except OCRUnavailable as exc:
        return jsonify({"success": False, "message": str(exc)}), 200
    except ExtractionError as exc:
        return jsonify({"success": False, "message": str(exc)}), 200
    except Exception as exc:  # noqa: BLE001
        return jsonify({"success": False, "message": f"Could not read file: {exc}"}), 200

    incoming = split_paragraphs(text)
    if not incoming:
        return jsonify({
            "success": False,
            "message": "No paragraphs could be extracted from that file.",
        }), 200

    data = _load_grouped(language)
    existing = data[language][difficulty]
    existing_set = {p.strip() for p in existing}

    added, duplicates = 0, 0
    for para in incoming:
        para = para.strip()
        if not para:
            continue
        if para in existing_set:
            duplicates += 1
            continue
        existing.append(para)
        existing_set.add(para)
        added += 1

    save_json(_lang_path(language), data)
    invalidate_paragraph_cache()

    return jsonify({
        "success": True,
        "added": added,
        "duplicates": duplicates,
        "total_found": len(incoming),
        "paragraphs": data[language],
    })


@admin.route("/paragraphs/clear", methods=["POST"])
@login_required
@admin_required
def clear_paragraphs():
    """Bulk delete every paragraph in a given language + difficulty."""
    payload = request.get_json(silent=True) or {}
    language = (payload.get("language") or "").lower()
    difficulty = (payload.get("difficulty") or "").lower()

    if language not in ("english", "hindi") or difficulty not in ("easy", "medium", "hard"):
        return jsonify({"success": False, "message": "Invalid request"}), 400

    data = _load_grouped(language)
    data[language][difficulty] = []
    save_json(_lang_path(language), data)
    invalidate_paragraph_cache()

    return jsonify({"success": True, "paragraphs": data[language]})


@admin.route("/paragraphs/export")
@login_required
@admin_required
def export_paragraphs():
    """Export a language's paragraphs as txt / json / csv / xlsx."""
    language = (request.args.get("language") or "english").lower()
    fmt = (request.args.get("format") or "txt").lower()

    if language not in ("english", "hindi"):
        language = "english"

    data = _load_grouped(language)[language]
    stamp = datetime.now().strftime("%Y%m%d")
    base = f"{language}_paragraphs_{stamp}"

    # Flatten into (difficulty, content) rows
    rows = []
    for diff in ("easy", "medium", "hard"):
        for content in data.get(diff, []):
            rows.append((diff, content))

    if fmt == "json":
        payload = json.dumps({language: data}, ensure_ascii=False, indent=2)
        return Response(
            payload,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={base}.json"},
        )

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["difficulty", "content"])
        writer.writerows(rows)
        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={base}.csv"},
        )

    if fmt in ("xlsx", "excel"):
        try:
            from openpyxl import Workbook
        except Exception:  # noqa: BLE001
            return jsonify({"success": False, "message": "Excel export unavailable."}), 500
        wb = Workbook()
        ws = wb.active
        ws.title = "Paragraphs"
        ws.append(["Difficulty", "Content"])
        for diff, content in rows:
            ws.append([diff, content])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"{base}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # default: txt (blank-line separated, difficulty headers)
    lines = []
    for diff in ("easy", "medium", "hard"):
        items = data.get(diff, [])
        if not items:
            continue
        lines.append(f"# {diff.upper()}")
        lines.append("")
        for content in items:
            lines.append(content)
            lines.append("")
    return Response(
        "\n".join(lines),
        mimetype="text/plain",
        headers={"Content-Disposition": f"attachment; filename={base}.txt"},
    )


@admin.route("/settings", methods=["GET", "POST"])
@login_required
@admin_required
def settings():

    if request.method == "POST":
        section = request.form.get("section", "general")
        updates = {}

        if section == "general":
            updates = {
                "site_name": request.form.get("site_name", "TypeMaster India").strip(),
                "default_timer": int(request.form.get("default_timer", 60)),
                "site_tagline": request.form.get("site_tagline", "").strip(),
                "maintenance_mode": request.form.get("maintenance_mode") == "on",
            }
        elif section == "branding":
            updates = {
                "org_name": request.form.get("org_name", "TypeMaster India").strip(),
                "founder_name": request.form.get("founder_name", "").strip(),
                "founder_title": request.form.get("founder_title", "Founder").strip(),
                "primary_color": request.form.get("primary_color", "#6366f1").strip(),
                "accent_color": request.form.get("accent_color", "#22d3ee").strip(),
            }
        elif section == "certificate":
            updates = {
                "certificate_prefix": request.form.get("certificate_prefix", "TMI").strip(),
                "certificate_footer": request.form.get("certificate_footer", "").strip(),
            }
        elif section == "email":
            updates = {
                "mail_from_name": request.form.get("mail_from_name", "").strip(),
                "mail_from_email": request.form.get("mail_from_email", "").strip(),
                "otp_enabled": request.form.get("otp_enabled") == "on",
            }
        elif section == "ocr":
            updates = {
                "ocr_default_lang": request.form.get("ocr_default_lang", "auto"),
                "ocr_deskew": request.form.get("ocr_deskew") == "on",
                "ocr_denoise": request.form.get("ocr_denoise") == "on",
            }
        elif section == "downloads":
            updates = {
                "downloads_enabled": request.form.get("downloads_enabled") == "on",
                "max_upload_mb": int(request.form.get("max_upload_mb", 50)),
            }
        elif section == "security":
            updates = {
                "require_email_verify": request.form.get("require_email_verify") == "on",
                "session_timeout_minutes": int(request.form.get("session_timeout_minutes", 0)),
            }

        if updates:
            save_settings(updates)
            flash("Settings saved.", "success")

        return redirect(url_for("admin.settings") + f"#{section}")

    settings_data = load_settings()

    return render_template(
        "admin/settings.html",
        settings=settings_data,
        ocr_ok=ocr_available(),
        mail_server=app_mail_config(),
    )


def app_mail_config():
    """Non-secret mail config for the settings UI."""
    from flask import current_app
    return {
        "server": current_app.config.get("MAIL_SERVER") or "Not configured",
        "port": current_app.config.get("MAIL_PORT", 587),
        "tls": current_app.config.get("MAIL_USE_TLS", True),
        "username": current_app.config.get("MAIL_USERNAME") or "Not configured",
    }


# ---------------------------------------------------------------------------
# Certificate Designer + Signature Manager
# ---------------------------------------------------------------------------

CERT_FONTS = [
    "Cinzel",
    "Playfair Display",
    "Cormorant Garamond",
    "Georgia",
    "Times New Roman",
]

CERT_POSITIONS = ["left", "center", "right"]
SIGNATURE_REL_PATH = "images/signature.png"
SIGNATURE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}


def _sample_cert_context():
    """Build sample data so the certificate preview renders without a real result."""
    sample_result = SimpleNamespace(
        gross_wpm=62,
        net_wpm=58,
        cpm=290,
        accuracy=97,
        errors=4,
        language="English",
        difficulty="Medium",
        duration=300,
        created_at=datetime.utcnow(),
    )
    sample_user = SimpleNamespace(full_name="Aarav Sharma")
    return sample_result, sample_user


@admin.route("/certificate")
@login_required
@admin_required
def certificate_designer():
    settings_data = load_settings()
    return render_template(
        "admin/certificate_designer.html",
        settings=settings_data,
        fonts=CERT_FONTS,
        positions=CERT_POSITIONS,
        has_signature=bool(signature_abs_path()),
        signature_url=(
            url_for("static", filename=SIGNATURE_REL_PATH) if signature_abs_path() else None
        ),
    )


@admin.route("/certificate", methods=["POST"])
@login_required
@admin_required
def save_certificate_design():
    font = request.form.get("cert_heading_font", "Cinzel").strip()
    qr_pos = request.form.get("cert_qr_position", "center").strip()
    sig_pos = request.form.get("cert_signature_position", "right").strip()

    updates = {
        "cert_title": request.form.get("cert_title", "").strip() or "Certificate of Achievement",
        "cert_subtitle": request.form.get("cert_subtitle", "").strip() or "Official Typing Performance Record",
        "cert_ribbon_enabled": request.form.get("cert_ribbon_enabled") == "on",
        "cert_ribbon_text": request.form.get("cert_ribbon_text", "").strip() or "CERTIFICATE OF EXCELLENCE",
        "cert_watermark_enabled": request.form.get("cert_watermark_enabled") == "on",
        "cert_watermark_text": request.form.get("cert_watermark_text", "").strip() or "TYPEMASTER INDIA",
        "cert_seal_enabled": request.form.get("cert_seal_enabled") == "on",
        "cert_logo_enabled": request.form.get("cert_logo_enabled") == "on",
        "cert_border_color": request.form.get("cert_border_color", "#0d3b8e").strip(),
        "cert_gold_color": request.form.get("cert_gold_color", "#c9a227").strip(),
        "cert_heading_font": font if font in CERT_FONTS else "Cinzel",
        "cert_qr_position": qr_pos if qr_pos in CERT_POSITIONS else "center",
        "cert_signature_position": sig_pos if sig_pos in CERT_POSITIONS else "right",
        "certificate_footer": request.form.get("certificate_footer", "").strip()
            or "Digitally generated and verifiable online.",
    }
    save_settings(updates)
    flash("Certificate design saved.", "success")
    return redirect(url_for("admin.certificate_designer"))


@admin.route("/certificate/preview")
@login_required
@admin_required
def certificate_preview():
    """Standalone certificate render used inside the designer's live preview iframe."""
    sample_result, sample_user = _sample_cert_context()
    return render_template(
        "admin/certificate_preview.html",
        result=sample_result,
        cert_user=sample_user,
        certificate_id="{}-2026-000123".format(load_settings().get("certificate_prefix", "TMI")),
        qr_code=None,
        verify_url="/verify/sample",
    )


@admin.route("/signature/upload", methods=["POST"])
@login_required
@admin_required
def upload_signature():
    file = request.files.get("signature")
    if not file or not file.filename:
        flash("Please choose a signature image to upload.", "danger")
        return redirect(url_for("admin.certificate_designer"))

    ext = os.path.splitext(secure_filename(file.filename))[1].lower()
    if ext not in SIGNATURE_EXTS:
        flash("Unsupported file type. Upload a PNG, JPG or WEBP image.", "danger")
        return redirect(url_for("admin.certificate_designer"))

    dest = os.path.join(BASE_DIR, "static", "images", "signature.png")
    os.makedirs(os.path.dirname(dest), exist_ok=True)

    make_transparent = request.form.get("make_transparent") == "on"

    try:
        from PIL import Image

        img = Image.open(file.stream).convert("RGBA")

        if make_transparent:
            # Turn near-white pixels transparent so the signature sits cleanly
            # on the certificate paper.
            datas = img.getdata()
            new_data = []
            for r, g, b, a in datas:
                if r > 235 and g > 235 and b > 235:
                    new_data.append((r, g, b, 0))
                else:
                    new_data.append((r, g, b, a))
            img.putdata(new_data)

        # Constrain very large uploads.
        max_w = 900
        if img.width > max_w:
            ratio = max_w / float(img.width)
            img = img.resize((max_w, int(img.height * ratio)), Image.LANCZOS)

        img.save(dest, "PNG")
    except Exception:
        # Fallback: store the raw upload as-is.
        file.stream.seek(0)
        file.save(dest)

    save_settings({"signature_path": SIGNATURE_REL_PATH})
    flash("Signature uploaded successfully. It now appears on all certificates.", "success")
    return redirect(url_for("admin.certificate_designer"))


@admin.route("/signature/delete", methods=["POST"])
@login_required
@admin_required
def delete_signature():
    path = signature_abs_path()
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            pass
        flash("Signature removed. Certificates will show the founder name only.", "success")
    else:
        flash("No signature file to remove.", "info")
    return redirect(url_for("admin.certificate_designer"))


@admin.route("/users/bulk-delete", methods=["POST"])
@login_required
@admin_required
def bulk_delete_users():
    ids = request.form.getlist("ids[]") or request.form.getlist("ids")
    deleted = 0
    for uid in ids:
        try:
            user = User.query.get(int(uid))
        except (TypeError, ValueError):
            continue
        if not user or user.id == current_user.id:
            continue
        db.session.delete(user)
        deleted += 1
    db.session.commit()
    flash(f"Deleted {deleted} user(s).", "success")
    return redirect(url_for("admin.users"))


@admin.route("/users/export")
@login_required
@admin_required
def export_users():
    fmt = request.args.get("format", "csv")
    users_list = User.query.order_by(User.id.desc()).all()
    stamp = datetime.now().strftime("%Y%m%d")

    if fmt in ("xlsx", "excel"):
        try:
            from openpyxl import Workbook
        except Exception:  # noqa: BLE001
            fmt = "csv"
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Full Name", "Username", "Email", "Role", "Level", "XP", "Joined"])
            for u in users_list:
                ws.append([
                    u.id, u.full_name, u.username, u.email,
                    "Admin" if u.is_admin else "User",
                    u.level, u.xp,
                    u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
                ])
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio, as_attachment=True,
                download_name=f"users_{stamp}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "full_name", "username", "email", "role", "level", "xp", "joined"])
    for u in users_list:
        writer.writerow([
            u.id, u.full_name, u.username, u.email,
            "Admin" if u.is_admin else "User",
            u.level, u.xp,
            u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
        ])
    return Response(
        buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=users_{stamp}.csv"},
    )


@admin.route("/results/export")
@login_required
@admin_required
def export_results():
    fmt = request.args.get("format", "csv")
    rows = Result.query.order_by(Result.created_at.desc()).all()
    stamp = datetime.now().strftime("%Y%m%d")

    if fmt in ("xlsx", "excel"):
        try:
            from openpyxl import Workbook
        except Exception:  # noqa: BLE001
            fmt = "csv"
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "User", "Language", "Difficulty", "Duration", "Net WPM", "Accuracy", "Errors", "Date"])
            for r in rows:
                ws.append([
                    r.id, r.user.username if r.user else "",
                    r.language, r.difficulty, r.duration,
                    r.net_wpm, r.accuracy, r.errors,
                    r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
                ])
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio, as_attachment=True,
                download_name=f"results_{stamp}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "user", "language", "difficulty", "duration", "net_wpm", "accuracy", "errors", "date"])
    for r in rows:
        writer.writerow([
            r.id, r.user.username if r.user else "",
            r.language, r.difficulty, r.duration,
            r.net_wpm, r.accuracy, r.errors,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        ])
    return Response(
        buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=results_{stamp}.csv"},
    )


@admin.route("/results/bulk-delete", methods=["POST"])
@login_required
@admin_required
def bulk_delete_results():
    ids = request.form.getlist("ids[]") or request.form.getlist("ids")
    deleted = 0
    for rid in ids:
        try:
            row = Result.query.get(int(rid))
        except (TypeError, ValueError):
            continue
        if row:
            db.session.delete(row)
            deleted += 1
    db.session.commit()
    flash(f"Deleted {deleted} result(s).", "success")
    return redirect(url_for("admin.results"))
